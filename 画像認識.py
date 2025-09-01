import os
import tkinter as tk
from tkinter import filedialog
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Tesseractのパスを指定

def select_image_and_process():
    """画像を選択し、OCR処理とExcelへの書き込みを行う関数"""
    
    # 1. 画像ファイル選択ダイアログを開く
    root = tk.Tk()
    root.withdraw()  # メインウィンドウを非表示にする
    file_path = filedialog.askopenfilename(
        title="画像ファイルを選択してください",
        filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")]
    )
    
    if not file_path:
        print("ファイルが選択されませんでした。")
        return

    try:
        # 2. 画像をメモリに読み込む
        print(f"画像ファイルを読み込み中: {file_path}")
        image = Image.open(file_path)

        # 3. OCR処理を実行し、テキストを抽出
        print("OCR処理を実行中...")
        extracted_text = pytesseract.image_to_string(image, lang='jpn+eng')
        
        # 4. 抽出したテキストをリストに分割
        # 空行や不要な空白を削除して、有効な文字列だけをリスト化
        processed_data = [line.strip() for line in extracted_text.splitlines() if line.strip()]

        if not processed_data:
            print("画像から文字が認識されませんでした。")
            return

        # 5. Excelファイルの準備
        excel_file = "ocr_results.xlsx"
        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
        else:
            workbook = openpyxl.Workbook()
            # デフォルトで作成される'Sheet'を削除
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        # 6. シート名の自動生成
        base_name = datetime.now().strftime("%m%d") # 例: 0830
        sheet_title = f"{base_name}(1)"
        counter = 1
        while sheet_title in workbook.sheetnames:
            counter += 1
            sheet_title = f"{base_name}({counter})"
        
        new_sheet = workbook.create_sheet(title=sheet_title)
        
        # 7. 抽出したデータをB1から横に書き込む
        print(f"Excelシート「{sheet_title}」にデータを書き込み中...")
        column_index = 2 # B列はインデックス2
        for data in processed_data:
            column_letter = get_column_letter(column_index)
            new_sheet[f'{column_letter}1'] = data
            column_index += 1
        
        # 8. Excelファイルを保存
        workbook.save(excel_file)
        print("Excelファイルへの書き込みが完了しました。")

    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    select_image_and_process()